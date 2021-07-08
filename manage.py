from parsers import gsi_parse_player_list, gsi_parse_stats, elo_str_to_elo_list, parse_payload_to_send, gsi_parse_names
from elo_calc import calc_elo_match
import server
import socket_client
import openpyxl
import config

GAMEMODE = config.GAMEMODE
STORAGE_TYPE = config.STORAGE_TYPE
FILENAME = config.FILENAME
SHEETNAME = config.SHEETNAME


def retrieve_current_elo_xlsx(FILENAME, SHEETNAME, players_str):
    wb = openpyxl.load_workbook(FILENAME)
    sheet = wb[SHEETNAME]

    players = players_str.split("/")

    current_elo = []
    row_counter = 1
    for player_index in range(len(players)):
        while True:
            if sheet.cell(row_counter, 1).value == players[player_index]:
                current_elo.append(int(sheet.cell(row_counter, 2).value()))
                break
            elif sheet.cell(row_counter, 1).value is None:
                return f"{players[player_index]} not found!"

    return current_elo



def retrieve_current_elo_socket(socket, players_in_match_str):
    """Sends request to server to send the players elo"""
    current_elo_str = socket_client.send_message(socket, players_in_match_str)
    #of the form "1505/1153/1539/1549"


def send_current_elo_xlsx(FILENAME, SHEETNAME, team_t, team_ct, new_elo_t, new_elo_ct):

    wb = openpyxl.load_workbook(FILENAME)
    sheet = wb[SHEETNAME]

    row_counter = 1
    for player_index in range(len(team_t)):
        if sheet.cell(row_counter, 1) == team_t[player_index]:
            sheet.cell(row_counter, 2).value = new_elo_t[player_index]

    row_counter = 1
    for player_index in range(len(team_ct)):
        if sheet.cell(row_counter, 1) == team_ct[player_index]:
            sheet.cell(row_counter, 2).value = new_elo_ct[player_index]

    wb.close()


def send_current_elo_socket(socket, updated_elo_str):
    socket_client.send_message(socket, updated_elo_str)


def exec_server(STORAGE_TYPE, GAMEMODE, custom_GAMEMODE, player_reference):

    """Starting Servers"""
    gsi_server_instance = server.GSIServer(("localhost",3000),"tau")
    gsi_server_instance.start_server()
    

    t_index = 0 #dont change
    ct_index = 1 #dont change
    max_team_wins = None
    if GAMEMODE == 1:
        max_team_wins = 16
    elif GAMEMODE == 2:
        max_team_wins = 9

    """Cheks if the round has ended in a T win (True, False), CT win (False, True), or draw (None, None)"""
    game_ended = (False, False)
    while game_ended[0] == False and game_ended[1] == False:
        
        game_ended = server.scan_for_win(gsi_server_instance, max_team_wins)
    
    gsi_server_output = server.output(gsi_server_instance)

    players_in_match_str = gsi_parse_player_list(gsi_server_output)
    # "Player1/Player2/Player3/Player4/Player5"

    # hint: variable names that contain "str" in them are usualy messages recieved from or ment to be send to the socket server 

    socket = None
    current_elo_str = ""
    if STORAGE_TYPE == 1:
        current_elo_str = retrieve_current_elo_xlsx(FILENAME, SHEETNAME, players_in_match_str)
        if isinstance(current_elo_str, str):
            print(current_elo_str)
            return

    elif STORAGE_TYPE == 2:
        socket = socket_client.start_client()
        current_elo_str = retrieve_current_elo_socket(socket, players_in_match_str)
    

    current_elo_list = elo_str_to_elo_list(current_elo_str, gsi_server_output)
    #of the form [[1505, 1153], [1539, 1549]]

    """Filters the players stats out of the gsi dictionary"""
    all_player_KAD = gsi_parse_stats(gsi_server_output)
    #of the form ([[22, 2, 16], [9, 4, 17]], [[18, 0, 16], [20, 4, 21]]) ([t1_stats,t2_stats],[ct1_stats,ct2_stats]) t1_stats = [kills,assists,deaths]


    """calculating the new elo"""
    new_elo = calc_elo_match(current_elo_list[t_index], current_elo_list[ct_index], game_ended[t_index], game_ended[ct_index], all_player_KAD[t_index], all_player_KAD[ct_index])


    """Parsing new elo to a string and sending it to the socket server"""
    updated_elo_str = parse_payload_to_send(new_elo,gsi_server_output)


    current_elo_str = ""
    if STORAGE_TYPE == 1:
        team_t, team_ct = gsi_parse_names(gsi_server_output)
        new_elo_t = new_elo[0]
        new_elo_ct = new_elo[1]
        current_elo_str = send_current_elo_xlsx(FILENAME, SHEETNAME, team_t, team_ct, new_elo_t, new_elo_ct)

    elif STORAGE_TYPE == 2:
        current_elo_str = send_current_elo_socket(socket, updated_elo_str)
    

exec_server()